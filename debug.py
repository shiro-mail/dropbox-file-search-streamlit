import streamlit as st
from dropbox_client import get_dropbox_client, get_files_in_folder
import xlrd
import openpyxl
import io

def debug_excel_file():
    """指定のExcelファイルの内容をデバッグ表示"""
    try:
        # Dropboxクライアントを取得
        dbx = get_dropbox_client()
        
        # まず、実際に存在するファイル一覧を表示
        st.write("=== 利用可能なファイル一覧 ===")
        files = get_files_in_folder("/恩藤さん")
        
        if not files:
            st.error("フォルダが見つからないか、ファイルがありません")
            return ""
        
        st.write(f"フォルダ内のファイル数: {len(files)}")
        for i, file in enumerate(files):
            st.write(f"{i+1}. {file['name']}")
        
        # 最初のExcelファイルを選択
        excel_files = [f for f in files if f['name'].lower().endswith(('.xls', '.xlsx'))]
        if not excel_files:
            st.error("Excelファイルが見つかりません")
            return ""
        
        # 最初のExcelファイルを処理
        target_file = excel_files[0]
        file_path = target_file['path']
        st.write(f"=== 処理対象ファイル: {file_path} ===")
        
        # ファイルをダウンロード
        st.write(f"=== ファイルダウンロード開始 ===")
        _, response = dbx.files_download(file_path)
        file_content = response.content
        st.write(f"ダウンロード完了: {len(file_content)} bytes")
        
        # Excelファイルの内容を抽出（形式に応じて分岐）
        st.write(f"=== Excelテキスト抽出開始 ===")
        
        text = ""
        if file_path.lower().endswith('.xlsx'):
            # .xlsxファイルはopenpyxlを使用
            st.write("openpyxlを使用して.xlsxファイルを処理")
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            st.write(f"ワークブック読み込み完了: {len(workbook.sheetnames)}シート")
            
            for sheet_name in workbook.sheetnames:
                st.write(f"シート処理中: {sheet_name}")
                sheet = workbook[sheet_name]
                st.write(f"シート情報: {sheet.max_row}行 x {sheet.max_column}列")
                
                text += f"シート: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text += row_text + "\n"
        else:
            # .xlsファイルはxlrdを使用
            st.write("xlrdを使用して.xlsファイルを処理")
            workbook = xlrd.open_workbook(file_contents=file_content)
            st.write(f"ワークブック読み込み完了: {workbook.nsheets}シート")
            
            for sheet_name in workbook.sheet_names():
                st.write(f"シート処理中: {sheet_name}")
                sheet = workbook.sheet_by_name(sheet_name)
                st.write(f"シート情報: {sheet.nrows}行 x {sheet.ncols}列")
                
                text += f"シート: {sheet_name}\n"
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        if cell_value:
                            row_data.append(str(cell_value))
                    if row_data:
                        text += " ".join(row_data) + "\n"
        
        st.write(f"=== 抽出完了: {len(text)}文字 ===")
        st.write("=== 抽出テキスト内容 ===")
        st.text_area("抽出テキスト", text, height=400)
        
        return text
        
    except Exception as e:
        st.error(f"!!! エラー発生: {e}")
        import traceback
        st.text(traceback.format_exc())
        return ""

# Streamlitアプリのメイン処理
st.title("Excelファイルデバッグ")
if st.button("デバッグ実行"):
    debug_excel_file()