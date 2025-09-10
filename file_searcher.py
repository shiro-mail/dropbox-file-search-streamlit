import os
import io
import PyPDF2
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None
try:
    from pdfminer.high_level import extract_text as pdfminer_extract_text
except Exception:
    pdfminer_extract_text = None
import docx
import openpyxl
import xlrd # .xlsファイル対応のために追加
from dropbox_client import get_dropbox_client, get_files_in_folder, get_subfolders 
from keyword_extractor import extract_keywords
# OCR 用ライブラリ（存在しなければ無視）
try:
    import pytesseract
    from PIL import Image
except Exception:
    pytesseract = None
    Image = None

# 再帰的にフォルダ配下のファイルを列挙

def _iter_files_recursive(root: str):
    stack = [root]
    while stack:
        cur = stack.pop()
        # ファイル
        for f in get_files_in_folder(cur):
            yield f
        # サブフォルダ
        try:
            subs = get_subfolders(cur) or []
        except Exception:
            subs = []
        for s in subs:
            p = s.get('full_path') or s.get('path')
            if p:
                stack.append(p)


def search_files(folder_path, user_input):
    """指定フォルダ内でファイルを検索（再帰）"""
    # キーワード抽出（関連度トップのみ）
    keywords = extract_keywords(user_input)
    
    if not keywords:
        return []

    # 除外記法のチェック
    exclude_keywords = [kw for kw in keywords if kw.get('type') == 'exclude']
    if exclude_keywords:
        return search_files_exclude(folder_path, exclude_keywords)
    
    # 関連度トップのキーワードを取得
    top_keyword = max(keywords, key=lambda x: x['relevance'])
    search_term = top_keyword['keyword']
    
    # 再帰でファイル一覧を取得
    files = list(_iter_files_recursive(folder_path))
    
    # ファイル名で検索
    search_results = []
    for file in files:
        if search_term.lower() in file['name'].lower():
            search_results.append({
                'file': file,
                'match_type': 'filename',
                'search_term': search_term
            })
    
    return search_results


def search_files_comprehensive(folder_path, user_input):
    """ファイル名と内容の両方で検索（再帰）"""
    # ファイル名検索
    filename_results = search_files(folder_path, user_input)
    
    # ファイル内容検索
    content_results = search_files_by_content(folder_path, user_input)
    
    # 結果を統合（重複除去）
    all_results = filename_results + content_results
    unique_results = []
    seen_files = set()
    
    for result in all_results:
        file_path = result['file']['path']
        if file_path not in seen_files:
            unique_results.append(result)
            seen_files.add(file_path)
    
    return unique_results


def search_files_by_content(folder_path, user_input):
    """ファイル内容で検索（再帰）"""
    # キーワード抽出（関連度トップのみ）
    keywords = extract_keywords(user_input)
    if not keywords:
        return []
    
    top_keyword = max(keywords, key=lambda x: x['relevance'])
    search_term = top_keyword['keyword']
    
    # 再帰でファイル一覧を取得
    files = list(_iter_files_recursive(folder_path))
    
    # ファイル内容で検索
    content_results = []
    for file in files:
        # ファイルをダウンロード
        file_content = download_file_content(file['path'])
        if file_content:
            # テキスト抽出（OCRフォールバック込み）
            text = extract_text_simple(file_content, file['name'])
            if search_term.lower() in text.lower():
                content_results.append({
                    'file': file,
                    'match_type': 'content',
                    'search_term': search_term
                })
    
    return content_results

def download_file_content(file_path):
    """ファイルをダウンロード"""
    try:
        dbx = get_dropbox_client()
        _, response = dbx.files_download(file_path)
        return response.content
    except Exception as e:
        return None

def extract_text_simple(file_content, filename):
    """ファイルの内容からテキストを抽出 (PDF, TXT, Excel, Word対応)"""
    try:
        text = ""
        _, file_ext =  os.path.splitext(filename)
        file_ext = file_ext.lower() # ← ここを修正

        if file_ext.endswith('.pdf'):
            # 1) PyMuPDF（文字化け耐性が高い）
            extracted = ""
            try:
                if fitz is not None:
                    with fitz.open(stream=file_content, filetype="pdf") as doc:
                        for p in doc:
                            extracted += p.get_text("text") or ""
            except Exception:
                extracted = extracted or ""

            # 2) PyPDF2 フォールバック
            if not extracted:
                try:
                    pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                    for page in pdf_reader.pages:
                        page_text = None
                        try:
                            page_text = page.extract_text()
                        except Exception:
                            page_text = None
                        if page_text:
                            extracted += page_text + "\n"
                except Exception:
                    pass

            # 3) pdfminer.six 最終フォールバック
            if not extracted and pdfminer_extract_text is not None:
                try:
                    extracted = pdfminer_extract_text(io.BytesIO(file_content)) or ""
                except Exception:
                    extracted = extracted or ""

            # 4) OCRフォールバック（画像PDF対策）
            if not extracted.strip() and pytesseract and fitz and Image:
                try:
                    with fitz.open(stream=file_content, filetype="pdf") as doc:
                        ocr_text = ""
                        # 最初の数ページのみOCR（処理負荷軽減）
                        for page_num in range(min(doc.page_count, 10)):
                            page = doc.load_page(page_num)
                            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                            # 日本語を優先（英数字混在の場合は 'jpn+eng' も可）
                            ocr_text += pytesseract.image_to_string(img, lang='jpn') + "\n"
                        if ocr_text.strip():
                            extracted = ocr_text
                except Exception:
                    pass

            text += extracted
            
        elif file_ext.endswith('.txt'):
            text = file_content.decode('utf-8', errors='ignore')
            
        elif file_ext.endswith('.docx'):
            doc = docx.Document(io.BytesIO(file_content))
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
        elif file_ext.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"シート: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text += row_text + "\n"
            
        elif file_ext.endswith('.xls'):
            workbook = xlrd.open_workbook(file_contents=file_content)
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                text += f"シート: {sheet_name}\n"
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        if cell_value:
                            row_data.append(str(cell_value))
                    if row_data:
                        text += " ".join(row_data) + "\n"
        else:
            return "" # 未対応のファイル形式は空文字列を返す
        
        return text
            
    except Exception as e:
        return ""


def search_files_exclude(folder_path, exclude_keywords):
    """除外記法でファイルを検索"""
    files = get_files_in_folder(folder_path)
    exclude_terms = [kw['keyword'][1:] for kw in exclude_keywords]  # "!"を除去
    
    search_results = []
    for file in files:
        should_exclude = False
        for term in exclude_terms:
            if term.startswith('.'):
                # 拡張子での除外
                if file['name'].lower().endswith(term.lower()):
                    should_exclude = True
                    break
            else:
                # ファイル名での除外
                if term.lower() in file['name'].lower():
                    should_exclude = True
                    break
        
        if not should_exclude:
            search_results.append({
                'file': file,
                'match_type': 'exclude_filter',
                'search_term': f"!{exclude_terms}"
            })
    
    return search_results